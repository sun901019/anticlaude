"""
從 蝦皮_選品-營運管理表_v5.xlsx 匯入初始資料進 DB
執行：python -m src.ecommerce.seed_from_excel
"""
import sys
import openpyxl
from pathlib import Path
from src.db.connection import db
from src.db.schema import init_db
from src.utils.logger import get_logger

log = get_logger("ecommerce.seed")

EXCEL_PATH = Path("蝦皮_選品-營運管理表_v5.xlsx")
EXCHANGE_RATE = 4.5
PLATFORM_FEE = 0.05
PAYMENT_FEE = 0.12


def _val(cell):
    return cell if cell is not None else None


def seed_products(ws_sku, ws_inventory, ws_perf):
    """從 01_選品資料庫 + 02_進貨紀錄 + 03_營運績效 匯入"""
    products = {}
    inventory_rows = []
    perf_rows = []

    # 01_選品資料庫 (跳過 header row 1)
    for row in ws_sku.iter_rows(min_row=2, values_only=True):
        sku = row[0]
        if not sku:
            continue
        cost_rmb = None
        # 從進貨紀錄取成本
        products[sku] = {
            "sku": str(sku),
            "name": str(row[1]) if row[1] else "",
            "product_type": str(row[2]) if row[2] else None,
            "scene": str(row[3]) if row[3] else None,
            "keyword": str(row[4]) if row[4] else None,
            "role": str(row[5]) if row[5] else None,
            "market_price_low": float(row[6]) if row[6] else None,
            "market_price_high": float(row[7]) if row[7] else None,
            "supplier": str(row[8]) if row[8] else None,
            "target_price": float(row[9]) if row[9] else None,
            "cost_rmb": None,
            "cost_twd": None,
        }

    # 02_進貨紀錄 (跳過 header)
    for row in ws_inventory.iter_rows(min_row=2, values_only=True):
        sku = row[0]
        if not sku or sku not in products:
            continue
        cost_rmb = float(row[3]) if row[3] else None
        cost_twd = float(row[5]) if row[5] else None
        qty = int(row[6]) if row[6] else 0
        lead_days = int(row[7]) if row[7] else None

        if cost_rmb:
            products[sku]["cost_rmb"] = cost_rmb
            products[sku]["cost_twd"] = cost_twd or (cost_rmb * EXCHANGE_RATE if cost_rmb else None)

        inventory_rows.append({
            "sku": str(sku),
            "purchase_date": str(row[2]) if row[2] else None,
            "cost_rmb": cost_rmb,
            "exchange_rate": float(row[4]) if row[4] else EXCHANGE_RATE,
            "cost_twd": cost_twd,
            "quantity": qty,
            "lead_days": lead_days,
            "supplier": str(row[8]) if row[8] else None,
        })

    # 03_營運績效 (跳過 header)
    for row in ws_perf.iter_rows(min_row=2, values_only=True):
        sku = row[0]
        if not sku or sku not in products:
            continue
        perf_rows.append({
            "sku": str(sku),
            "record_date": "2026-03-14",
            "current_price": float(row[2]) if row[2] else None,
            "sales_7d": int(row[3]) if row[3] else 0,
            "revenue_7d": float(row[4]) if row[4] else 0,
            "ad_spend_7d": float(row[5]) if row[5] else 0,
            "gross_profit": float(row[6]) if row[6] else None,
            "gross_margin": float(row[7]) if row[7] else None,
            "current_stock": int(row[8]) if row[8] else 0,
            "roas": float(row[9]) if row[9] else None,
            "next_action": str(row[10]) if row[10] else None,
        })

    return products, inventory_rows, perf_rows


def run():
    init_db()

    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws_sku = wb["01_選品資料庫"]
    ws_inv = wb["02_進貨紀錄"]
    ws_perf = wb["03_營運績效與毛利"]

    products, inventory_rows, perf_rows = seed_products(ws_sku, ws_inv, ws_perf)

    with db() as conn:
        # 匯入商品主檔
        for p in products.values():
            conn.execute("""
                INSERT OR REPLACE INTO fl_products
                (sku, name, product_type, scene, keyword, role,
                 market_price_low, market_price_high, supplier,
                 cost_rmb, cost_twd, target_price)
                VALUES (:sku, :name, :product_type, :scene, :keyword, :role,
                        :market_price_low, :market_price_high, :supplier,
                        :cost_rmb, :cost_twd, :target_price)
            """, p)
        log.info(f"匯入商品主檔：{len(products)} 筆")

        # 匯入進貨紀錄
        for inv in inventory_rows:
            if inv["quantity"] and inv["quantity"] > 0:
                conn.execute("""
                    INSERT INTO fl_inventory
                    (sku, purchase_date, cost_rmb, exchange_rate, cost_twd, quantity, lead_days, supplier)
                    VALUES (:sku, :purchase_date, :cost_rmb, :exchange_rate, :cost_twd, :quantity, :lead_days, :supplier)
                """, inv)
        log.info(f"匯入進貨紀錄：{len(inventory_rows)} 筆")

        # 匯入績效數據
        for perf in perf_rows:
            conn.execute("""
                INSERT OR REPLACE INTO fl_performance
                (sku, record_date, current_price, sales_7d, revenue_7d,
                 ad_spend_7d, gross_profit, gross_margin, current_stock, roas, next_action)
                VALUES (:sku, :record_date, :current_price, :sales_7d, :revenue_7d,
                        :ad_spend_7d, :gross_profit, :gross_margin, :current_stock, :roas, :next_action)
            """, perf)
        log.info(f"匯入績效數據：{len(perf_rows)} 筆")

    print(f"✓ 匯入完成：{len(products)} 個 SKU")


if __name__ == "__main__":
    run()
